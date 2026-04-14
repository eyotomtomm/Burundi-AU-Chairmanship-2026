"""
Report generation utilities for the Diplomatic Portal admin.
Generates Excel (openpyxl) and PDF (reportlab) reports for
users, events, articles, and full analytics.
"""

import io
from datetime import timedelta

from django.contrib.auth.models import User
from django.db.models import Count, Sum, Q
from django.utils import timezone

# ──────────────────────────────────────────────────────────────
# Excel report helpers
# ──────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _style_header_row(ws, num_cols):
    """Apply bold white-on-slate header to the first row."""
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        bottom=Side(style='thin', color='CBD5E1'),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[1].height = 30


def _auto_width(ws):
    """Auto-adjust column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted = min(max_length + 4, 50)
        ws.column_dimensions[col_letter].width = adjusted


def _add_summary_row(ws, row_num, label, value, num_cols):
    """Add a bold summary row at the bottom of a sheet."""
    summary_font = Font(name='Arial', bold=True, size=11)
    summary_fill = PatternFill(start_color='F1F5F9', end_color='F1F5F9', fill_type='solid')
    cell = ws.cell(row=row_num, column=1, value=label)
    cell.font = summary_font
    cell.fill = summary_fill
    cell = ws.cell(row=row_num, column=2, value=value)
    cell.font = summary_font
    cell.fill = summary_fill
    for col in range(3, num_cols + 1):
        ws.cell(row=row_num, column=col).fill = summary_fill


# ──────────────────────────────────────────────────────
# Excel: Users Report
# ──────────────────────────────────────────────────────
def generate_users_excel(queryset, filters=None):
    """Generate an Excel workbook with user data.
    Returns a BytesIO buffer.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Users'

    columns = [
        'ID', 'Username', 'Email', 'First Name', 'Last Name',
        'Nationality', 'Badge', 'Verified', 'Active',
        'Date Joined', 'Last Login',
    ]
    ws.append(columns)
    _style_header_row(ws, len(columns))

    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    row_idx = 2
    for user in queryset.select_related('profile').order_by('-date_joined'):
        profile = getattr(user, 'profile', None)
        row = [
            user.id,
            user.username,
            user.email,
            user.first_name,
            user.last_name,
            profile.get_nationality_display() if profile and profile.nationality else '',
            profile.badge_type if profile else '',
            'Yes' if (profile and profile.is_verified) else 'No',
            'Yes' if user.is_active else 'No',
            user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else '',
            user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never',
        ]
        ws.append(row)
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = data_font
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        row_idx += 1

    # Summary row
    total_row = row_idx
    verified_count = sum(1 for u in queryset.select_related('profile')
                         if getattr(u, 'profile', None) and u.profile.is_verified)
    _add_summary_row(ws, total_row, f'Total: {queryset.count()} users', f'Verified: {verified_count}', len(columns))

    # Filters info sheet
    if filters:
        ws2 = wb.create_sheet(title='Filters Applied')
        ws2.append(['Filter', 'Value'])
        _style_header_row(ws2, 2)
        for k, v in filters.items():
            ws2.append([k, str(v)])
        _auto_width(ws2)

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────
# Excel: Events Report
# ──────────────────────────────────────────────────────
def generate_events_excel(queryset):
    """Generate an Excel workbook with event data."""
    from core.models import EventSubmission

    wb = Workbook()
    ws = wb.active
    ws.title = 'Events'

    columns = ['ID', 'Name', 'Date', 'Location', 'Registrations', 'Status']
    ws.append(columns)
    _style_header_row(ws, len(columns))

    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    status_green = Font(name='Arial', size=10, color='16A34A')
    status_gray = Font(name='Arial', size=10, color='64748B')

    row_idx = 2
    for event in queryset.order_by('-event_date'):
        now = timezone.now()
        status = 'Active' if event.is_active and event.event_date >= now else (
            'Inactive' if not event.is_active else 'Past'
        )
        reg_count = EventSubmission.objects.filter(
            registration__event_name__icontains=event.name[:30]
        ).count() if hasattr(event, 'name') else 0

        row = [
            event.id,
            event.name,
            event.event_date.strftime('%Y-%m-%d %H:%M') if event.event_date else '',
            event.address,
            reg_count,
            status,
        ]
        ws.append(row)
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = data_font
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        # Color-code status
        status_cell = ws.cell(row=row_idx, column=6)
        if status == 'Active':
            status_cell.font = status_green
        else:
            status_cell.font = status_gray
        row_idx += 1

    _add_summary_row(ws, row_idx, f'Total: {queryset.count()} events', '', len(columns))
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────
# Excel: Articles Report
# ──────────────────────────────────────────────────────
def generate_articles_excel(queryset):
    """Generate an Excel workbook with article data."""
    from core.models import ArticleComment

    wb = Workbook()
    ws = wb.active
    ws.title = 'Articles'

    columns = ['ID', 'Title', 'Author', 'Category', 'Publish Date', 'Views', 'Likes', 'Comments']
    ws.append(columns)
    _style_header_row(ws, len(columns))

    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')

    row_idx = 2
    for article in queryset.select_related('category').order_by('-publish_date'):
        comment_count = ArticleComment.objects.filter(article=article).count()
        row = [
            article.id,
            article.title,
            article.author,
            article.category.name if article.category else '',
            article.publish_date.strftime('%Y-%m-%d %H:%M') if article.publish_date else '',
            article.view_count,
            article.like_count,
            comment_count,
        ]
        ws.append(row)
        for col in range(1, len(columns) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = data_font
            if row_idx % 2 == 0:
                cell.fill = alt_fill
        row_idx += 1

    # Summary row
    total_views = queryset.aggregate(s=Sum('view_count'))['s'] or 0
    total_likes = queryset.aggregate(s=Sum('like_count'))['s'] or 0
    _add_summary_row(ws, row_idx, f'Total: {queryset.count()} articles',
                     f'Views: {total_views}  |  Likes: {total_likes}', len(columns))
    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────
# Excel: Full Analytics Report (multi-sheet)
# ──────────────────────────────────────────────────────
def generate_analytics_excel(stats_dict):
    """Generate a multi-sheet analytics Excel workbook.
    stats_dict should contain keys: users, content, engagement, support.
    """
    wb = Workbook()

    # ── Sheet 1: Users ──
    ws = wb.active
    ws.title = 'Users'
    user_rows = [
        ['Metric', 'Value'],
        ['Total Users', stats_dict.get('total_users', 0)],
        ['New Users (7 days)', stats_dict.get('new_7d', 0)],
        ['New Users (30 days)', stats_dict.get('new_30d', 0)],
        ['Active Today', stats_dict.get('active_today', 0)],
        ['Active (7 days)', stats_dict.get('active_7d', 0)],
        ['Active (30 days)', stats_dict.get('active_30d', 0)],
        ['Verified Users', stats_dict.get('verified_users', 0)],
    ]
    for row in user_rows:
        ws.append(row)
    _style_header_row(ws, 2)
    _auto_width(ws)

    # ── Sheet 2: Content ──
    ws2 = wb.create_sheet(title='Content')
    content_rows = [
        ['Content Type', 'Total', 'Views', 'Likes'],
        ['Articles', stats_dict.get('total_articles', 0),
         stats_dict.get('article_views', 0), stats_dict.get('article_likes', 0)],
        ['Magazines', stats_dict.get('total_magazines', 0),
         stats_dict.get('magazine_views', 0), stats_dict.get('magazine_likes', 0)],
        ['Videos', stats_dict.get('total_videos', 0),
         stats_dict.get('video_views', 0), stats_dict.get('video_likes', 0)],
        ['Gallery Albums', stats_dict.get('total_albums', 0),
         stats_dict.get('album_views', 0), stats_dict.get('album_likes', 0)],
    ]
    for row in content_rows:
        ws2.append(row)
    _style_header_row(ws2, 4)
    _auto_width(ws2)

    # ── Sheet 3: Engagement ──
    ws3 = wb.create_sheet(title='Engagement')
    engagement_rows = [
        ['Metric', 'Value'],
        ['Total Event Submissions', stats_dict.get('total_event_submissions', 0)],
        ['Total Check-ins', stats_dict.get('total_checkins', 0)],
        ['Total Polls', stats_dict.get('total_polls', 0)],
        ['Total Discussions', stats_dict.get('total_discussions', 0)],
    ]
    for row in engagement_rows:
        ws3.append(row)
    _style_header_row(ws3, 2)
    _auto_width(ws3)

    # ── Sheet 4: Support ──
    ws4 = wb.create_sheet(title='Support')
    support_rows = [
        ['Metric', 'Value'],
        ['Open Tickets', stats_dict.get('open_tickets', 0)],
        ['In-Progress Tickets', stats_dict.get('in_progress_tickets', 0)],
        ['Resolved (30 days)', stats_dict.get('tickets_resolved_30d', 0)],
        ['Total Tickets', stats_dict.get('total_tickets', 0)],
    ]
    for row in support_rows:
        ws4.append(row)
    _style_header_row(ws4, 2)
    _auto_width(ws4)

    # ── Sheet 5: Top Nationalities ──
    if stats_dict.get('nationality_data'):
        ws5 = wb.create_sheet(title='Nationalities')
        ws5.append(['Country', 'Users'])
        _style_header_row(ws5, 2)
        for entry in stats_dict['nationality_data']:
            ws5.append([entry.get('nationality', ''), entry.get('count', 0)])
        _auto_width(ws5)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────────────
# PDF report helpers
# ──────────────────────────────────────────────────────────────
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable,
)
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.graphics.charts.barcharts import VerticalBarChart


def _build_header(title, subtitle=None):
    """Build a list of flowables for the report header."""
    styles = getSampleStyleSheet()
    header_style = ParagraphStyle(
        'ReportHeader',
        parent=styles['Title'],
        fontSize=22,
        textColor=colors.HexColor('#1E293B'),
        spaceAfter=4,
        fontName='Helvetica-Bold',
    )
    sub_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=12,
    )
    date_style = ParagraphStyle(
        'ReportDate',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#94A3B8'),
        spaceAfter=6,
    )

    elements = []
    # App branding line
    elements.append(Paragraph('Burundi Chairmanship 2026-2027', sub_style))
    elements.append(Paragraph(title, header_style))
    if subtitle:
        elements.append(Paragraph(subtitle, sub_style))
    elements.append(Paragraph(
        f'Generated on {timezone.now().strftime("%B %d, %Y at %H:%M UTC")}',
        date_style,
    ))
    elements.append(HRFlowable(
        width='100%', thickness=1,
        color=colors.HexColor('#E2E8F0'),
        spaceAfter=20,
    ))
    return elements


def _build_section_title(text):
    """Build a styled section heading."""
    style = ParagraphStyle(
        'SectionTitle',
        fontSize=14,
        textColor=colors.HexColor('#1E293B'),
        spaceBefore=16,
        spaceAfter=8,
        fontName='Helvetica-Bold',
    )
    return Paragraph(text, style)


def _build_table(headers, data_rows, col_widths=None):
    """Build a styled table with alternating row colors."""
    table_data = [headers] + data_rows

    style_cmds = [
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        # Data rows
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    # Alternating row colors
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(
                ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F8FAFC'))
            )

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle(style_cmds))
    return table


def _build_bar_chart(labels, values, title='', width=450, height=200):
    """Build a simple vertical bar chart drawing."""
    if not values or all(v == 0 for v in values):
        return Spacer(1, 10)

    drawing = Drawing(width, height + 40)

    # Title
    drawing.add(String(10, height + 20, title,
                       fontName='Helvetica-Bold', fontSize=10,
                       fillColor=colors.HexColor('#1E293B')))

    chart = VerticalBarChart()
    chart.x = 40
    chart.y = 20
    chart.width = width - 80
    chart.height = height - 30
    chart.data = [values]
    chart.categoryAxis.categoryNames = labels
    chart.categoryAxis.labels.fontSize = 7
    chart.categoryAxis.labels.angle = 30
    chart.valueAxis.valueMin = 0
    chart.valueAxis.labels.fontSize = 7
    chart.bars[0].fillColor = colors.HexColor('#D97706')
    chart.bars[0].strokeColor = colors.HexColor('#B45309')
    chart.bars[0].strokeWidth = 0.5

    drawing.add(chart)
    return drawing


def _page_footer(canvas, doc):
    """Add page number footer."""
    canvas.saveState()
    canvas.setFont('Helvetica', 7)
    canvas.setFillColor(colors.HexColor('#94A3B8'))
    canvas.drawCentredString(
        A4[0] / 2, 15 * mm,
        f'Burundi Chairmanship Diplomatic Portal  |  Page {doc.page}'
    )
    canvas.restoreState()


# ──────────────────────────────────────────────────────
# PDF: Full Analytics Report
# ──────────────────────────────────────────────────────
def generate_analytics_pdf(stats_dict):
    """Generate a comprehensive analytics PDF report.
    Returns a BytesIO buffer.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=25 * mm, bottomMargin=25 * mm,
        leftMargin=20 * mm, rightMargin=20 * mm,
        title='Analytics Report - Burundi Chairmanship',
    )

    elements = []

    # ── Header ──
    elements.extend(_build_header(
        'Analytics Report',
        'Comprehensive platform performance overview',
    ))

    # ── Executive Summary ──
    elements.append(_build_section_title('Executive Summary'))
    styles = getSampleStyleSheet()
    body = styles['Normal']
    body.fontSize = 9
    body.textColor = colors.HexColor('#334155')
    body.leading = 14

    total_users = stats_dict.get('total_users', 0)
    active_30d = stats_dict.get('active_30d', 0)
    total_articles = stats_dict.get('total_articles', 0)

    summary_text = (
        f'The platform currently serves <b>{total_users}</b> registered users, '
        f'with <b>{active_30d}</b> active in the last 30 days. '
        f'There are <b>{total_articles}</b> published articles, '
        f'<b>{stats_dict.get("total_events", 0)}</b> events, and '
        f'<b>{stats_dict.get("total_magazines", 0)}</b> magazine editions. '
        f'This report provides a detailed breakdown of user engagement, '
        f'content performance, and support metrics.'
    )
    elements.append(Paragraph(summary_text, body))
    elements.append(Spacer(1, 16))

    # ── User Statistics Table ──
    elements.append(_build_section_title('User Statistics'))
    user_table_data = [
        ['Total Users', str(total_users)],
        ['New Users (7 days)', str(stats_dict.get('new_7d', 0))],
        ['New Users (30 days)', str(stats_dict.get('new_30d', 0))],
        ['Active Today', str(stats_dict.get('active_today', 0))],
        ['Active (7 days)', str(stats_dict.get('active_7d', 0))],
        ['Active (30 days)', str(active_30d)],
        ['Verified Users', str(stats_dict.get('verified_users', 0))],
    ]
    elements.append(_build_table(['Metric', 'Value'], user_table_data,
                                 col_widths=[250, 200]))
    elements.append(Spacer(1, 16))

    # ── Content Performance Table ──
    elements.append(_build_section_title('Content Performance'))
    content_table_data = [
        ['Articles', str(stats_dict.get('total_articles', 0)),
         str(stats_dict.get('article_views', 0)), str(stats_dict.get('article_likes', 0))],
        ['Magazines', str(stats_dict.get('total_magazines', 0)),
         str(stats_dict.get('magazine_views', 0)), str(stats_dict.get('magazine_likes', 0))],
        ['Videos', str(stats_dict.get('total_videos', 0)),
         str(stats_dict.get('video_views', 0)), str(stats_dict.get('video_likes', 0))],
        ['Gallery', str(stats_dict.get('total_albums', 0)),
         str(stats_dict.get('album_views', 0)), str(stats_dict.get('album_likes', 0))],
    ]
    elements.append(_build_table(
        ['Type', 'Total', 'Views', 'Likes'],
        content_table_data,
        col_widths=[130, 100, 100, 100],
    ))
    elements.append(Spacer(1, 12))

    # Content bar chart
    content_labels = ['Articles', 'Magazines', 'Videos', 'Gallery']
    content_views = [
        stats_dict.get('article_views', 0),
        stats_dict.get('magazine_views', 0),
        stats_dict.get('video_views', 0),
        stats_dict.get('album_views', 0),
    ]
    elements.append(_build_bar_chart(content_labels, content_views,
                                     title='Content Views by Type'))
    elements.append(Spacer(1, 12))

    # ── Event Analytics ──
    elements.append(PageBreak())
    elements.append(_build_section_title('Event Analytics'))
    event_table_data = [
        ['Total Events', str(stats_dict.get('total_events', 0))],
        ['Active Events', str(stats_dict.get('active_events', 0))],
        ['Event Submissions', str(stats_dict.get('total_event_submissions', 0))],
        ['Check-ins', str(stats_dict.get('total_checkins', 0))],
    ]
    elements.append(_build_table(['Metric', 'Value'], event_table_data,
                                 col_widths=[250, 200]))
    elements.append(Spacer(1, 16))

    # ── Support Metrics ──
    elements.append(_build_section_title('Support Metrics'))
    support_table_data = [
        ['Open Tickets', str(stats_dict.get('open_tickets', 0))],
        ['In Progress', str(stats_dict.get('in_progress_tickets', 0))],
        ['Resolved (30 days)', str(stats_dict.get('tickets_resolved_30d', 0))],
        ['Total Tickets', str(stats_dict.get('total_tickets', 0))],
    ]
    elements.append(_build_table(['Metric', 'Value'], support_table_data,
                                 col_widths=[250, 200]))
    elements.append(Spacer(1, 16))

    # ── Nationality Distribution ──
    nat_data = stats_dict.get('nationality_data', [])
    if nat_data:
        elements.append(_build_section_title('Top Nationalities'))
        nat_table_rows = [
            [entry.get('nationality', ''), str(entry.get('count', 0))]
            for entry in nat_data[:15]
        ]
        elements.append(_build_table(['Country', 'Users'], nat_table_rows,
                                     col_widths=[300, 150]))
        elements.append(Spacer(1, 12))

        nat_labels = [e.get('nationality', '')[:12] for e in nat_data[:10]]
        nat_values = [e.get('count', 0) for e in nat_data[:10]]
        elements.append(_build_bar_chart(nat_labels, nat_values,
                                         title='Top 10 Nationalities'))

    # ── Engagement Summary ──
    elements.append(PageBreak())
    elements.append(_build_section_title('Engagement Summary'))
    engagement_table_data = [
        ['Polls Created', str(stats_dict.get('total_polls', 0))],
        ['Discussions', str(stats_dict.get('total_discussions', 0))],
        ['Active Announcements', str(stats_dict.get('active_announcements', 0))],
        ['Resources Published', str(stats_dict.get('total_resources', 0))],
    ]
    elements.append(_build_table(['Metric', 'Value'], engagement_table_data,
                                 col_widths=[250, 200]))

    doc.build(elements, onFirstPage=_page_footer, onLaterPages=_page_footer)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────────────
# PDF: Single User Profile Report
# ──────────────────────────────────────────────────────
def generate_user_report_pdf(user):
    """Generate a PDF report for a single user profile.
    Returns a BytesIO buffer.
    """
    from core.models import ArticleLike, ArticleComment, SupportTicket, EventSubmission

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=25 * mm, bottomMargin=25 * mm,
        leftMargin=20 * mm, rightMargin=20 * mm,
        title=f'User Report - {user.username}',
    )

    elements = []
    profile = getattr(user, 'profile', None)

    # Header
    elements.extend(_build_header(
        f'User Report: {user.get_full_name() or user.username}',
        f'Account ID: {user.id}',
    ))

    # ── Account Details ──
    elements.append(_build_section_title('Account Details'))
    details = [
        ['Username', user.username],
        ['Email', user.email or 'Not provided'],
        ['Full Name', user.get_full_name() or 'Not provided'],
        ['Date Joined', user.date_joined.strftime('%B %d, %Y') if user.date_joined else ''],
        ['Last Login', user.last_login.strftime('%B %d, %Y %H:%M') if user.last_login else 'Never'],
        ['Account Active', 'Yes' if user.is_active else 'No'],
        ['Staff Status', 'Yes' if user.is_staff else 'No'],
    ]
    elements.append(_build_table(['Field', 'Value'], details,
                                 col_widths=[200, 250]))
    elements.append(Spacer(1, 12))

    # ── Profile Information ──
    if profile:
        elements.append(_build_section_title('Profile Information'))
        profile_details = [
            ['Nationality', profile.get_nationality_display() if profile.nationality else 'Not set'],
            ['Gender', profile.get_gender_display() if profile.gender else 'Not set'],
            ['Phone', profile.phone_number or 'Not provided'],
            ['Language', profile.get_preferred_language_display()],
            ['Badge Type', profile.badge_type or 'None'],
            ['Verified', 'Yes' if profile.is_verified else 'No'],
            ['Email Verified', 'Yes' if profile.is_email_verified else 'No'],
            ['Device', profile.device_type or 'Unknown'],
            ['OS', profile.device_os or 'Unknown'],
            ['App Version', profile.app_version or 'Unknown'],
            ['Account Deactivated', 'Yes' if profile.is_deactivated else 'No'],
        ]
        elements.append(_build_table(['Field', 'Value'], profile_details,
                                     col_widths=[200, 250]))
        elements.append(Spacer(1, 12))

    # ── Activity Summary ──
    elements.append(_build_section_title('Activity Summary'))
    article_likes = ArticleLike.objects.filter(user=user).count()
    article_comments = ArticleComment.objects.filter(user=user).count()
    tickets = SupportTicket.objects.filter(user=user).count()
    event_subs = EventSubmission.objects.filter(user=user).count()

    activity = [
        ['Article Likes', str(article_likes)],
        ['Article Comments', str(article_comments)],
        ['Event Registrations', str(event_subs)],
        ['Support Tickets', str(tickets)],
    ]
    elements.append(_build_table(['Activity', 'Count'], activity,
                                 col_widths=[250, 200]))

    doc.build(elements, onFirstPage=_page_footer, onLaterPages=_page_footer)
    buf.seek(0)
    return buf
